"""
build_historical.py
===================
Genera data/historical.csv a partir de los ECO xlsx que existan en data/.

Ejecutar UNA VEZ (o cada vez que agregues nuevos ECO xlsx):
    python build_historical.py

El CSV resultante se usa para entrenar XGBoost al iniciar app.py.

Columnas de salida (9 columnas):
    anio, mes, ipc_factor, descripcion, unidad, grupo_unidad,
    cantidad, precio_unitario, total
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

# ── Configuración ──────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_DIR  = BASE_DIR / 'data'
IPC_CSV   = DATA_DIR / 'IPC_VAR_MEN1_HIST_NEW.csv'
OUT_CSV   = DATA_DIR / 'historical.csv'

# Patrón de nombre de archivo ECO
# Ej: A22M451_FLESAN_ECO01_2026-02-11_ADJUDICADO.xlsx
ECO_RE = re.compile(
    r'(?P<lic>[A-Z0-9]+)_(?P<empresa>[A-Z0-9]+)_ECO\d+_'
    r'(?P<anio>\d{4})-(?P<mes>\d{2})',
    re.IGNORECASE,
)

# Mapa de grupos de unidad — idéntico al de xgboost_estimator.py
GRUPO_UNIDAD: dict[str, str] = {
    'hh': 'mano_obra',  'hr': 'mano_obra',  'hrs': 'mano_obra',
    'hm': 'hora_maquina',
    'dia': 'dia',        'día': 'dia',        'jornada': 'dia',
    'm3': 'volumetrico',
    'ml': 'lineal',      'lm': 'lineal',      'm': 'lineal',
    'm2': 'area',
    'kg': 'peso',        'ton': 'peso',
    'gl': 'global',      'glb': 'global',
    'un': 'unidad',      'cu': 'unidad',      'und': 'unidad',
}
GRUPO_DEFAULT = 'otros'

# ── IPC ────────────────────────────────────────────────────────────────────────
_MES_MAP = {
    'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
    'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12,'sept':9,
}


def cargar_ipc(ruta: Path) -> dict[tuple[int,int], float]:
    """Lee el CSV del Banco Central. Devuelve {(año, mes): ipc_acumulado}."""
    if not ruta.exists():
        print(f"[IPC] ADVERTENCIA — no se encontró {ruta}. Se usará ipc_factor=1.0 para todos.")
        return {}
    try:
        df = pd.read_csv(ruta, sep=';', skiprows=2, header=None)
        df.columns = ['Periodo','IPC_var','IPCX','IPCX1','IPC_SAE']
        df = df.dropna(subset=['Periodo'])

        def parse_period(p):
            partes = str(p).split('.')
            if len(partes) != 2:
                return None, None
            mes_str, year_str = partes
            return int(year_str), _MES_MAP.get(mes_str.lower().strip())

        df[['Year','Month']] = df['Periodo'].apply(lambda x: pd.Series(parse_period(x)))
        df = df.dropna(subset=['Year','Month'])
        df['IPC_var'] = pd.to_numeric(
            df['IPC_var'].astype(str).str.replace(',','.'), errors='coerce'
        )
        df['IPC_acc'] = 100 * (1 + df['IPC_var'] / 100).cumprod()
        return {(int(r['Year']), int(r['Month'])): float(r['IPC_acc'])
                for _, r in df.iterrows() if pd.notna(r['IPC_acc'])}
    except Exception as e:
        print(f"[IPC] Error al cargar: {e}")
        return {}


def ipc_para_fecha(anio: int, mes: int, ipc_dict: dict) -> float:
    """Devuelve el IPC acumulado del período, buscando ±6 meses si falta."""
    v = ipc_dict.get((anio, mes))
    if v:
        return v
    for delta in range(1, 7):
        for signo in (-1, 1):
            m2 = mes + signo * delta
            a2 = anio + (m2 - 1) // 12
            m2 = ((m2 - 1) % 12) + 1
            v = ipc_dict.get((a2, m2))
            if v:
                return v
    return list(ipc_dict.values())[-1] if ipc_dict else 100.0


# ── Parser de ECO xlsx ─────────────────────────────────────────────────────────

def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f > 0 else None
    s = re.sub(r'[$\s,]', '', str(v).strip())
    try:
        f = float(s)
        return f if f > 0 else None
    except ValueError:
        return None


def _is_item_code(v) -> bool:
    return bool(v and re.match(r'^\d+(\.\d+)*$', str(v).strip()))


def extraer_items_xlsx(ruta: Path) -> list[dict]:
    """Extrae ítems de un ECO xlsx.

    Detecta automáticamente si es formato simple (sintético) o complejo (FLESAN real).
    """
    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
    except Exception as e:
        print(f"  [SKIP] {ruta.name} — no se pudo abrir: {e}")
        return []

    ws = wb.active

    # ── Buscar fila de encabezado ────────────────────────────────────────────
    header_row = None
    header_cols: dict[str, int] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 0):
        row_lower = [str(v).strip().lower() if v is not None else '' for v in row]
        if any('ítem' in s or s == 'item' for s in row_lower):
            header_row = r_idx
            for c_idx, s in enumerate(row_lower):
                if 'ítem' in s or s == 'item':
                    header_cols['item'] = c_idx
                elif 'descripción' in s or 'descripcion' in s or s == 'nombre':
                    header_cols['descripcion'] = c_idx
                elif 'unidad' in s:
                    header_cols['unidad'] = c_idx
                elif 'cantidad' in s:
                    header_cols['cantidad'] = c_idx
                elif 'precio unitario' in s or 'precio_unit' in s:
                    header_cols['precio_unitario'] = c_idx
            break
  

    items = []

    # ── Formato simple (encabezado detectado) ───────────────────────────────
    header_row = None
    header_cols: dict[str, tuple[int, int]] = {} # Ahora guarda límites (inicio, fin)

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 0):
        row_lower = [str(v).strip().lower() if v is not None else '' for v in row]
        if any('ítem' in s or s == 'item' for s in row_lower):
            header_row = r_idx
            
            # 1. Detectar en qué índices exactos hay texto (los títulos)
            indices_texto = [i for i, val in enumerate(row_lower) if val != '']
            
            # 2. Asignar el "bloque" (corralito) a cada columna
            for i, c_idx in enumerate(indices_texto):
                s = row_lower[c_idx]
                
                # El límite derecho es donde empieza el siguiente título (o el fin de la fila)
                limite_derecho = indices_texto[i + 1] if i + 1 < len(indices_texto) else len(row_lower)
                
                if 'ítem' in s or s == 'item':
                    header_cols['item'] = (c_idx, limite_derecho)
                elif 'descripción' in s or 'descripcion' in s or s == 'nombre':
                    header_cols['descripcion'] = (c_idx, limite_derecho)
                elif 'unidad' in s:
                    header_cols['unidad'] = (c_idx, limite_derecho)
                elif 'cantidad' in s:
                    header_cols['cantidad'] = (c_idx, limite_derecho)
                elif 'precio unitario' in s or 'precio_unit' in s or 'precio' in s:
                    header_cols['precio_unitario'] = (c_idx, limite_derecho)
                elif 'total' in s:
                    header_cols['total'] = (c_idx, limite_derecho)
            break

    items = []

    # ── Formato simple (encabezado detectado) ───────────────────────────────
    if header_row is not None and len(header_cols) >= 4:
        
        # Función que extrae el primer dato que encuentre dentro de sus propios límites
        def _extraer_del_bloque(fila, rango_columnas):
            if not rango_columnas:
                return None
            inicio, fin = rango_columnas
            for i in range(inicio, min(fin, len(fila))):
                valor = fila[i]
                if valor is not None and str(valor).strip() != '':
                    return valor
            return None

        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            # Usamos los límites que calculamos arriba
            c_item = _extraer_del_bloque(row, header_cols.get('item'))
            c_desc = _extraer_del_bloque(row, header_cols.get('descripcion'))
            c_unit = _extraer_del_bloque(row, header_cols.get('unidad'))
            c_qty  = _extraer_del_bloque(row, header_cols.get('cantidad'))
            c_pu   = _extraer_del_bloque(row, header_cols.get('precio_unitario'))

            if not _is_item_code(c_item):
                continue
                
            precio   = _to_float(c_pu)
            cantidad = _to_float(c_qty)
            unidad   = str(c_unit).strip() if c_unit else ''

            if precio and precio > 100 and cantidad and unidad:
                items.append({
                    'descripcion':     str(c_desc).strip() if c_desc else '',
                    'unidad':          unidad,
                    'cantidad':        cantidad,
                    'precio_unitario': precio,
                })
        return items

    # ── Formato complejo FLESAN (sin encabezado claro) ───────────────────────
    # item=col0, desc=col2, unit=col12, qty=col15, pu=col18
    for row in ws.iter_rows(values_only=True):
        if len(row) < 19:
            continue
        c_item = row[0]
        c_desc = row[2]
        c_unit = row[12]
        c_qty  = row[15]
        c_pu   = row[18]

        if not _is_item_code(c_item):
            continue
        precio   = _to_float(c_pu)
        cantidad = _to_float(c_qty)
        unidad   = str(c_unit).strip() if c_unit else ''

        if precio and precio > 500 and cantidad and unidad not in ['', '-', 'nan']:
            items.append({
                'descripcion':     str(c_desc).strip() if c_desc else '',
                'unidad':          unidad,
                'cantidad':        cantidad,
                'precio_unitario': precio,
            })

    return items


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  build_historical.py — generando data/historical.csv")
    print("=" * 65)

    ipc_dict = cargar_ipc(IPC_CSV)
    ipc_ref  = max(ipc_dict.values()) if ipc_dict else 100.0
    print(f"[IPC] {len(ipc_dict)} meses cargados. Referencia más reciente: {ipc_ref:.2f}\n")

    eco_files = sorted([
        f for f in DATA_DIR.glob('*_ECO*.xlsx')
        if f.stat().st_size > 1_000
    ])

    if not eco_files:
        print(f"ERROR — No se encontraron *_ECO*_ADJUDICADO.xlsx en {DATA_DIR}")
        sys.exit(1)

    print(f"Encontrados {len(eco_files)} archivos ECO:\n")

    all_rows: list[dict] = []

    for eco_path in eco_files:
        
        m = ECO_RE.search(eco_path.stem)
        if not m:
            print(f"  [SKIP] {eco_path.name} — no coincide con patrón de nombre")
            continue

        anio = int(m.group('anio'))
        mes  = int(m.group('mes'))



        # --- NUEVO: Calcular peso de la muestra ---
        nombre_archivo = eco_path.stem.upper()
        if 'ADJUDICADO' in nombre_archivo:
            peso = 1.0
        else:
            peso = 0.3  # Puedes ajustar este valor entre 0.1 y 0.9 según prefieras
        # -----------------------------------------

        ipc_abs    = ipc_para_fecha(anio, mes, ipc_dict)
        ipc_factor = round(ipc_abs / ipc_ref, 6)

        items = extraer_items_xlsx(eco_path)

        for it in items:
            u = it['unidad'].lower().strip()
            all_rows.append({
                'anio':            anio,
                'mes':             mes,
                'ipc_factor':      ipc_factor,
                'peso_muestra':    peso, # <--- NUEVO
                'descripcion':     it['descripcion'],
                'unidad':          it['unidad'],
                'grupo_unidad':    GRUPO_UNIDAD.get(u, GRUPO_DEFAULT),
                'cantidad':        it['cantidad'],
                'precio_unitario': it['precio_unitario'],
                'total':           round(it['precio_unitario'] * it['cantidad'], 2),
            })

        print(
            f"  ✓ {eco_path.name:<55} {anio}-{mes:02d}  "
            f"→ {len(items):>3} ítems  ipc_factor={ipc_factor:.4f}"
        )




    if not all_rows:
        print("\nERROR — No se extrajo ningún ítem. Verifica el formato de los xlsx.")
        sys.exit(1)

    df = pd.DataFrame(all_rows, columns=[
        'anio', 'mes', 'ipc_factor', 'peso_muestra', # <--- NUEVO
        'descripcion', 'unidad', 'grupo_unidad',
        'cantidad', 'precio_unitario', 'total',
    ])
    df.to_csv(OUT_CSV, index=False, encoding='utf-8-sig')

    # ── Distribución por grupo ────────────────────────────────────────────────
    total = len(df)
    print(f"\n{'='*65}")
    print(f"  historical.csv generado — {total} filas\n")
    print("[Historical] Distribución por grupo_unidad:")
    for grupo, cnt in sorted(df['grupo_unidad'].value_counts().items(), key=lambda x: -x[1]):
        pct = 100 * cnt / total
        print(f"  {grupo:<15} → {cnt:>5} registros ({pct:.1f}%)")

    print(
        f"\n  P.U.: min=${df['precio_unitario'].min():,.0f}  "
        f"max=${df['precio_unitario'].max():,.0f}  "
        f"mediana=${df['precio_unitario'].median():,.0f}"
    )
    print(f"  Guardado en: {OUT_CSV}")
    print(f"{'='*65}")
    print("\n  Listo. Reinicia app.py para que XGBoost se entrene con estos datos.\n")


if __name__ == '__main__':
    main()
